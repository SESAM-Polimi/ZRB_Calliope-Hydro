# -*- coding: utf-8 -*-
"""
Created on Tue Dec  3 11:04:23 2019

@authors: Nicolò Stevanato - Alessandro Barbieri
            Politecnico di Milano - Department of Energy
          Martina Daddi
            Politecnico di Milano - Department of Electronics, Information and Bioengineering
"""
import openpyxl
import calliope
import os
import configparser
import time
import shutil
from plotting_fausto import plotting
import datetime
import sys
import platform


if __name__ == "__main__":
    start = time.time()
    print(f'Starting at: {time.strftime("%H:%M:%S")}.\n')

    # Settings loading
    config = configparser.ConfigParser(allow_no_value=False, inline_comment_prefixes='#')
    config.read_file(open("Calliope_Hydro_config.conf"))
    # Bools
    debug = config.getboolean('Settings', 'debug')
    spillage_fausto = config.getboolean('Settings', 'spillage_fausto')
    plotting_fa = config.getboolean('Settings', 'plotting_fa')
    restore_original_timeseries = config.getboolean('Settings', 'restore_original_timeseries')
    plot = config.getboolean('Settings', 'plot')
    summary = config.getboolean('Settings', 'summary')
    save = config.getboolean('Settings', 'save')
    save_settings = config.getboolean('Settings', 'save_settings')
    # Strings
    verbosity = config.get('Settings', 'verbosity')
    # info_string = config.get('Settings', 'info_about_this_run')
    # Ints
    spillage_percentage = config.getfloat('Settings', 'spillage_percentage')
    iterations = config.getint('Settings', 'iterations')

    if restore_original_timeseries:
        shutil.copytree("Timeseries/Initialize_Loop", "Timeseries", dirs_exist_ok=True)
    calliope.set_log_verbosity(verbosity,  include_solver_output=False)

    if debug:
        iterations = 1
        results_dir_general = f"Results/results_{time.strftime('%Y%m%d-%H%M%S')}_debugmode"
    elif not spillage_fausto:
        results_dir_general = f"Results/results_{time.strftime('%Y%m%d-%H%M%S')}_nospillage"
    else:
        results_dir_general = f"Results/results_{time.strftime('%Y%m%d-%H%M%S')}"

    for iteration in range(iterations):
        print(f'\n\nIteration {iteration + 1} of {iterations} (time: {time.strftime("%H:%M:%S")}).\n')

        # check for convergence if iteration != 0
        #

        results_folder = f"{results_dir_general}/results_{iteration + 1}"
        if save and iteration != 0:  # nella prima iterazione copio il contenuto di Initaialize Loop in Timeseries, quindi non serve sapere quali siano gli input (sono quelli di Ini.Loop)
            folder_timeseries = f'{results_folder}/input_timeseries'
            os.makedirs(folder_timeseries)
            for file in ['effCB.txt', 'effITT.txt', 'effKG.txt', 'effKA.txt',
                         'spillageA_eff.txt', 'spillageB_eff.txt', 'spillageC_eff.txt', 'spillageD_eff.txt',
                         'evapLoss_CB.txt', 'evapLoss_ITT.txt', 'evapLoss_KA.txt', 'evapLoss_KG.txt']:
                shutil.copy(f'Timeseries/{file}', folder_timeseries)

        if debug:
            model = calliope.read_netcdf('Results/results_debug/results_1/model.nc')
        else:
            model = calliope.Model('model.yaml')
            model.run()

        if save:
            model.to_csv(f'{results_folder}/results_csv')
            model.to_netcdf(f"{results_folder}/model.nc")

        if plotting_fa:
            plots_folder = f"{results_folder}/plots"
            plotting.FaPlotting(model).write_excel(path=plots_folder, exist_ok=False)
            plotting.FaPlotting(model).plot_storage_timeseries(path=plots_folder, spillage_coeff=spillage_percentage)
            plotting.FaPlotting(model).plot_storage_and_carriers(path=plots_folder, spillage_coeff=spillage_percentage)

        if not debug and iterations != 1 and iteration != iterations - 1:
            # create identical dataframes for the new variables that I'm going to create to be updated in cleaner code
            DataFrame8 = model.get_formatted_array('carrier_prod').loc[{'techs': 'PV', 'carriers': 'power', 'locs': ['Zambia']}].to_pandas().T  # FIXME: ci sono molti più valori in questi df rispetto ai 17520 aggiornati
            DataFrame13 = model.get_formatted_array('carrier_prod').loc[{'techs': 'OCGT', 'carriers': 'power', 'locs': ['Moz-North-Center']}].to_pandas().T

            # ITT
            # extract storage value, initialize efficiency, downstream water to storage, basin surface and evaporation losses
            # upload EXCEL file with evaporation rates

            StorageA_nuovo = model.get_formatted_array('storage').loc[{'techs': 'storageA', 'locs': ['Zambia']}].to_pandas().T  # extract results from the calliope-model
            eff_conv_ITT = DataFrame8.copy()
            evapLoss_ITT = DataFrame8.copy()
            supStorage_ITT = DataFrame8.copy()
            if spillage_fausto:
                spillageA_eff = DataFrame8.copy()
            Data_ITT = openpyxl.load_workbook('Timeseries/lsv_min_max_rel_ITT.xlsx')  # upload excel with ITT data
            EvapRate_ITT = Data_ITT['evap+salto']  # select the sheet with rvaporation data

            # cycle to calculate every timestep's values of efficiency and other initialized variables
            for a in range(0, 17520):
                eff_conv_ITT.iloc[a] = (40.5-(1030.5-(-5*10**(-19)*(StorageA_nuovo.iloc[a] + 699000000)**2 + 8*10**(-9)*(StorageA_nuovo.iloc[a] + 699000000)+1000.7)))*9.8*1000*0.9/3600*10**(-3)
                supStorage_ITT.iloc[a] = -3*10**(-12)*(StorageA_nuovo.iloc[a] + 699000000)**2 + 0.08*(StorageA_nuovo.iloc[a] + 699000000)+3*10**7
                b = a + 2
                evapLoss_ITT.iloc[a] = (10**-2) * abs((supStorage_ITT.iloc[a]*EvapRate_ITT.cell(row=b, column=6).value/1000/EvapRate_ITT.cell(row=b, column=2).value/24)/(StorageA_nuovo.iloc[a] + 699000000)) # TODO: calcolare valori reali! non '(10**-2) * abs()' a caso
                if spillage_fausto:
                    if float(StorageA_nuovo.iloc[a]) > spillage_percentage * model._model_data.data_vars['storage_cap'].loc['Zambia::storageA']:
                        spillageA_eff.iloc[a] = 1
                    else:
                        spillageA_eff.iloc[a] = 0

            # remove old timeseries, replace with new ones
            os.remove('Timeseries/effITT.txt')
            os.remove('Timeseries/evapLoss_ITT.txt')
            eff_conv_ITT.to_csv('Timeseries/effITT.txt')
            evapLoss_ITT.to_csv('Timeseries/evapLoss_ITT.txt')
            if spillage_fausto:
                os.remove('Timeseries/spillageA_eff.txt')
                spillageA_eff.to_csv('Timeseries/spillageA_eff.txt')
            # Kafue Gorge
            StorageB_nuovo = model.get_formatted_array('storage').loc[{'techs': 'storageB','locs':['Zambia']}].to_pandas().T  # extract results from the calliope-model
            eff_conv_KG = DataFrame8.copy()
            evapLoss_KG = DataFrame8.copy()
            supStorage_KG = DataFrame8.copy()
            if spillage_fausto:
                spillageB_eff = DataFrame8.copy()
            Data_KG = openpyxl.load_workbook('Timeseries/lsv_min_max_rel_KGU.xlsx')  # upload excel with KGU data
            EvapRate_KG = Data_KG['evap+salto']  # select the sheet with rvaporation data

            # cycle to calculate every timestep's values of efficiency and other initialized variables
            for k in range(0, 17520):
                eff_conv_KG.iloc[k]=(397-(977.6-(957*(StorageB_nuovo.iloc[k] + 5000000)**0.001)))*9.8*1000*0.9/3600*10**(-3)
                supStorage_KG.iloc[k]= -10**(-10)*(StorageB_nuovo.iloc[k] + 5000000)**2 + 1.129*(StorageB_nuovo.iloc[k] + 5000000)-10**7
                h = k + 2
                evapLoss_KG.iloc[k]= (10**-2) * abs((supStorage_KG.iloc[k]*EvapRate_KG.cell(row=h, column=6).value/1000/EvapRate_KG.cell(row=h,column=2).value/24)/(StorageB_nuovo.iloc[k] + 5000000))
                if spillage_fausto:
                    if float(StorageB_nuovo.iloc[k]) > spillage_percentage * model._model_data.data_vars['storage_cap'].loc['Zambia::storageB']:
                        spillageB_eff.iloc[k] = 1
                    else:
                        spillageB_eff.iloc[k] = 0
            # remove old timeseries, replace with new ones
            os.remove('Timeseries/effKG.txt')
            os.remove('Timeseries/evapLoss_KG.txt')
            eff_conv_KG.to_csv('Timeseries/effKG.txt')
            evapLoss_KG.to_csv('Timeseries/evapLoss_KG.txt')
            if spillage_fausto:
                os.remove('Timeseries/spillageB_eff.txt')
                spillageB_eff.to_csv('Timeseries/spillageB_eff.txt')

            #Kariba
            StorageC_nuovo = model.get_formatted_array('storage').loc[{'techs': 'storageC','locs':['Zambia']}].to_pandas().T  # extract results from the calliope-model
            eff_conv_KA = DataFrame8.copy()
            evapLoss_KA = DataFrame8.copy()
            supStorage_KA = DataFrame8.copy()
            if spillage_fausto:
                spillageC_eff = DataFrame8.copy()
            Data_KA=openpyxl.load_workbook('Timeseries/lsv_min_max_rel_KA.xlsx')  # upload excel with KA data
            EvapRate_KA=Data_KA['evap+salto']  # select the sheet with rvaporation data

            # cycle to calculate every timestep's values of efficiency and other initialized variables
            for j in range(0, 17520):
                eff_conv_KA.iloc[j] = (110-(489.5-(-5*10**(-23)*(StorageC_nuovo.iloc[j] + 116054000000)**2 + 2*10**(-10)*(StorageC_nuovo.iloc[j] + 116054000000)+452.89)))*9.8*1000*0.9/3600*10**(-3)
                supStorage_KA.iloc[j] = -10**(-13)*(StorageC_nuovo.iloc[j] + 116054000000)**2 + 0.0493*(StorageC_nuovo.iloc[j] + 116054000000)+4*10**6
                l = j + 2
                evapLoss_KA.iloc[j] = (10**-2) * abs((supStorage_KA.iloc[j]*EvapRate_KA.cell(row=l,column=6).value/1000/EvapRate_KA.cell(row=l,column=2).value/24)/(StorageC_nuovo.iloc[j] + 116054000000))
                if spillage_fausto:
                    if float(StorageC_nuovo.iloc[j]) > spillage_percentage * model._model_data.data_vars['storage_cap'].loc['Zambia::storageC']:
                        spillageC_eff.iloc[j] = 1
                    else:
                        spillageC_eff.iloc[j] = 0

            # remove old timeseries, replace with new ones
            os.remove('Timeseries/effKA.txt')
            os.remove('Timeseries/evapLoss_KA.txt')
            eff_conv_KA.to_csv('Timeseries/effKA.txt')
            evapLoss_KA.to_csv('Timeseries/evapLoss_KA.txt')
            if spillage_fausto:
                os.remove('Timeseries/spillageC_eff.txt')
                spillageC_eff.to_csv('Timeseries/spillageC_eff.txt')

            #Cahora
            StorageD_nuovo=model.get_formatted_array('storage').loc[{'techs': 'storageD','locs':['Moz-North-Center']}].to_pandas().T # extract results from the calliope-model
            eff_conv_CB = DataFrame13.copy()
            evapLoss_CB = DataFrame13.copy()
            supStorage_CB = DataFrame13.copy()
            if spillage_fausto:
                spillageD_eff = DataFrame13.copy()
            Data_CB=openpyxl.load_workbook('Timeseries/lsv_min_max_rel_CB.xlsx') # upload excel with CB data
            EvapRate_CB=Data_CB['evap+salto'] # select the sheet with rvaporation data

            # cycle to calculate every timestep's values of efficiency and other initialized variables
            for i in range(0, 17520):
                eff_conv_CB.iloc[i]=(128-(331-(6*10**(-21)*(StorageD_nuovo.iloc[i] + 32000000)**2 + 9*10**(-10)*(StorageD_nuovo.iloc[i] + 32000000)+294.16)))*9.8*1000*0.9/3600*10**(-3)  #riempo variabile eff con eff calcolata in funzione dello storage
                supStorage_CB.iloc[i]= -2*10**(-13)*(StorageD_nuovo.iloc[i] + 32000000)**2 + 0.0469*(StorageD_nuovo.iloc[i] + 32000000)+8*10**8
                c=i+2
                evapLoss_CB.iloc[i]= (10**-2) * abs((supStorage_CB.iloc[i]*EvapRate_CB.cell(row=c,column=6).value/1000/EvapRate_CB.cell(row=c,column=2).value/24)/(StorageD_nuovo.iloc[i] + 32000000))
                if spillage_fausto:
                    if float(StorageD_nuovo.iloc[i]) > spillage_percentage * model._model_data.data_vars['storage_cap'].loc['Moz-North-Center::storageD']:
                        spillageD_eff.iloc[i] = 1
                    else:
                        spillageD_eff.iloc[i] = 0

            # remove old timeseries, replace with new ones
            os.remove('Timeseries/effCB.txt')
            os.remove('Timeseries/evapLoss_CB.txt')
            eff_conv_CB.to_csv('Timeseries/effCB.txt')
            evapLoss_CB.to_csv('Timeseries/evapLoss_CB.txt')
            if spillage_fausto:
                os.remove('Timeseries/spillageD_eff.txt')
                spillageD_eff.to_csv('Timeseries/spillageD_eff.txt')

        # Trying to plot with calliope's plot methods
        if plot:
            try:
                if summary:
                    model.plot.summary(to_file=f'{results_folder}/PLOT modello_{iteration + 1}.html')  # o '.svg' per avere file vettoriali
                    # model.plot.timeseries(to_file=f'{directory_scelta}/{nome_cartella_salvataggio}/PLOT modello_{ii}.svg')
                else:
                    model.plot.timeseries(to_file=f'{results_folder}/PLOT modello_{iteration + 1}.svg')
                    model.plot.transmission()
                    # model.plot.timeseries(subset={'techs': ['ccgt', 'battery', 'csp']})  # esempio NATIONAL SCALE
                    # model.plot.flows() # ci mette un po'
            except Exception as exc:
                print(f'Calliope plotting error: {exc}.')

    elapsed = time.time() - start
    if save_settings:
        with open('Calliope_Hydro_config.conf') as file:
            lines = [line.split('#')[0].rstrip() for line in file if line.rstrip() != '']
        other_lines = f'\n[Other Infos]' \
                      f"\n- name of the results' directory:{results_dir_general} " \
                      f'\n- Time spent to run the script (days, h:m:s): {str(datetime.timedelta(seconds=elapsed))}' \
                      f'\n- Calliope_Hydro.py was executed with "PyCharm debugger" ' \
                      f'(if "False" the script was just "run", without the debugger): ' \
                      f'{True if sys.gettrace() else False}' \
                      f'\n- System which run the script: {platform.uname()}' \
                      f'\n- (Guess on the system which run the script: '
        nl = '\n'
        with open(f'{results_dir_general}/Settings.txt', 'w') as outfile:
            outfile.write(f'{nl.join(lines)}{other_lines}')
